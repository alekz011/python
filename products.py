#! /usr/bin/python3

# Corrects costs in produce sales spreadsheet

import openpyxl


wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The prduce types and their updated prices
PRICE_UPDATES = { 'Garlic': 3.07, 
                  'Celey' : 1.19,
                  'Lemon': 1.27}

# Loop through the rows and update the prices
for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
